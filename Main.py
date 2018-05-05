import os
import time
import datetime
import random
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl import Workbook
from selenium import webdriver

data_xls_poz_map = {
    1: {'name': '风险等级', 'code': 'risk_level'},
    2: {'name': '赔标/信用标', 'code': 'pei'},
    3: {'name': '完成度', 'code': 'progress'},
    4: {'name': '用户名', 'code': 'user_name'},
    5: {'name': '信用等级', 'code': 'credit_level'},
    6: {'name': '贷款金额', 'code': 'amount'},
    7: {'name': '利率', 'code': 'rate'},
    8: {'name': '还款期限', 'code': 'term'},
    9: {'name': '性别', 'code': 'sex'},
    10: {'name': '年龄', 'code': 'age'},
    11: {'name': '文化程度', 'code': 'edu_bg'},
    12: {'name': '学习形式', 'code': 'learn_way'},
    13: {'name': '借款用途', 'code': 'lend_purpose'},
    14: {'name': '还款来源', 'code': 'payment'},
    15: {'name': '工作信息', 'code': 'work_info'},
    16: {'name': '收入情况', 'code': 'income_info'},
    17: {'name': '认证状况', 'code': 'verfied_info'},
    18: {'name': '成功借款次数', 'code': 'sucuess_cnt'},
    19: {'name': '成功还款次数', 'code': 'sucuess_repayment_cnt'},
    20: {'name': '正常还清次数', 'code': 'normal_repayment_cnt'},
    21: {'name': '历史记录', 'code': 'history_info'},
    22: {'name': '逾期(0-15天)还清次数', 'code': 'delay_lt15_repayment_cnt'},
    23: {'name': '逾期(15天以上)还清次数', 'code': 'delay_gt15_repayment_cnt'},
    24: {'name': '累计借贷金额', 'code': 'amount_sum'},
    25: {'name': '待还金额', 'code': 'unreturned_amount'},
    26: {'name': '待收金额', 'code': 'unreceived_amount'},
    27: {'name': '历史最高负债', 'code': 'biggest_lend_amount'},
    28: {'name': '单笔最高借款金额', 'code': 'biggest_debt_amount'}
}

TYPE_KPT = 4  # LoanCategoryId 4:平衡型,8:保守型,5:进取型
TYPE_SORT = 1  # 0不排序,1降序,2升序
file_name = 'data.xlsx'  # 存储数据文件名
today = datetime.date.today()  # 启动date
now = datetime.datetime.now()  # 启动datetime
TYPE_KPT_MAP = {4: '平衡型', 8: '保守型', 5: '进取型'}  # 类型映射Map
row = 2  # 表格数据开始行

login_url = 'https://ac.ppdai.com/User/Login'  # 登陆链接

browser = webdriver.Firefox()  # 使用火狐浏览器


# html转换
def html_to_soup(url):
    browser.get(url)
    return BeautifulSoup(browser.page_source, 'xml')


# url构造器 获得爬取链接
def url_constructor(page_index, _type):
    base_url = 'https://invest.ppdai.com/loan/' \
               'listnew?LoanCategoryId=' + str(_type) + \
               '&PageIndex=' + str(page_index) + \
               '&SortType=' + str(TYPE_SORT) + '&MinAmount=0&MaxAmount=0'
    aim_url = base_url
    return aim_url


# 获取链接页面所包含的详情页链接并封装到List
def details_url_list_getter(url):
    soup = html_to_soup(url)
    # 整个信息列表所在的div
    div_outer_borrow_list = soup.find('div', attrs={'class': 'outerBorrowList'})
    if '很抱歉，热门列表已经被抢空啦' in str(div_outer_borrow_list):
        return []
    else:
        # 详情信息所在的a标签
        a_s = div_outer_borrow_list.find_all('a', attrs={'class': 'title ell'})
        details_urls = []
        for a in a_s:
            details_urls.append(a['href'])
        return details_urls


# 详情页信息提取
def details_info_getter(details_url):
    try:  # 增加异常捕获，避免不可预见错误导致程序崩溃
        soup = html_to_soup(details_url)

        # 满标页面 (图片alt内容)
        if '借款成功' in soup:
            print('满标页面....')

        # 整个信息列表所在的div
        wrap_new_lend_detail = soup.find('div', attrs={'class': 'wrapNewLendDetailInfoLeft'})

        # 包含信用等级的a [信用等级详细值在其中span标签的class属性中]
        credit_level = wrap_new_lend_detail.find('a', attrs={'class': 'altQust'}).find('span')['class'] \
            .replace('creditRating ', '')

        # 用户名
        user_name = wrap_new_lend_detail.find('span', attrs={'class': 'username'}).get_text()

        # 赔标/信用标
        pei_i = soup.find('i', attrs={'class': 'pei'})

        # 完成度
        progress = soup.find('div', attrs={'class': 'newLendDetailRefundLeft'}) \
            .find('div', attrs={'class': 'part clearfix'}) \
            .find('div').get_text().replace('进度条：', '').strip()

        # 借款的详细信息 [借款金额 , 协议利率 , 期限]
        new_lend_detail_money_list = soup.find('div', attrs={'class': 'newLendDetailMoneyLeft'}).find_all('dd')

        # 人员详细信息 [性别 , 年龄 , 注册时间 , 文化程度 , 毕业院校 , 学习形式 , 借款用途 , 还款来源 , 工作信息 , 收入情况]
        lender_info_list = soup.find('div', attrs={'class': 'lender-info'}).findAll('p', attrs={'class': 'ex col-1'})

        # 信息可能不完全存在 , 需要特殊处理
        sex = '--'  # 性别
        age = '--'  # 年龄
        regist_date = '--'  # 注册时间
        edu_bg = '--'  # 文化程度
        graduate = '--'  # 毕业院校
        learn_way = '--'  # 学习形式
        lend_purpose = '--'  # 借款用途
        payment = '--'  # 还款来源
        work_info = '--'  # 工作信息
        income_info = '--'  # 收入情况
        for it in lender_info_list:
            if '性别' in str(it):
                sex = it.find('span').get_text()
            elif '年龄' in str(it):
                age = it.find('span').get_text()
            elif '注册时间' in str(it):
                regist_date = it.find('span').get_text()
            elif '文化程度' in str(it):
                edu_bg = it.find('span').get_text()
            elif '毕业院校' in str(it):
                graduate = it.find('span').get_text()
            elif '学习形式' in str(it):
                learn_way = it.find('span').get_text()
            elif '借款用途' in str(it):
                lend_purpose = it.find('span').get_text()
            elif '还款来源' in str(it):
                payment = it.find('span').get_text()
            elif '工作信息' in str(it):
                work_info = it.find('span').get_text()
            elif '收入情况' in str(it):
                income_info = it.find('span').get_text()

        # 认证信息 [可能不存在]
        record_info_list_parent = soup.find('ul', attrs={'class', 'record-info'})
        record_info_list = []
        if record_info_list_parent is not None:
            record_info_list = record_info_list_parent.findAll('li')
        verfied_info = ''
        for i in range(len(record_info_list)):
            verfied_info += (record_info_list[i].get_text()) + ' '
        verfied_info = '--' if (not verfied_info) else verfied_info.strip()  # 内容为空时默认写入 `--`

        # 总体div
        tab_contain_divs = soup.find("div", attrs={'class': 'lendDetailTab_tabContent w1000center'}) \
            .find_all('div', attrs={'class': 'tab-contain'})

        # 统计信息
        statistics_info_list = ['', '', '', '', '', '', '', '', '', '', '', '']
        for content in tab_contain_divs:  # Fixed:完成页面不包含统计信息
            if "统计信息" in str(content):
                statistics_info_list = content.find_all('span', attrs={'class', 'num'})
        # 投资人情况List
        ol_list_parent = soup.find('div', attrs={'class': 'scroll-area'})
        ol_list = []
        if ol_list_parent is not None:
            ol_list = ol_list_parent.find_all('ol')
        else:
            print('当前页面暂无投资人信息')

        # 投资人信息列表
        investor_list = []

        # 投资人单条信息
        for ol in ol_list:
            # 单条投资人信息
            investor_info = {}
            li_list = ol.find_all('li')
            investor_info['investor_id'] = li_list[0].find('span', attrs={'class': 'listname'}).get_text()  # 投资人id
            investment_way_tag = li_list[0].find('a', attrs={'target': '_blank'})  # 投资方式 [根据图标区分 , 可能没有]
            if 'title' in str(investment_way_tag):
                investor_info['investment_way'] = investment_way_tag['title']
            else:
                investor_info['investment_way'] = '未知'

            investor_info['rate'] = li_list[1].get_text()  # 利率
            investor_info['term'] = li_list[2].get_text()  # 期限
            investor_info['valid_amount'] = li_list[3].get_text().replace('¥', '')  # 有效金额
            investor_info['investment_date'] = li_list[4].get_text()  # 投标时间
            investor_list.append(investor_info)

        # 将信息放入字典中
        result_dic = {}  # 返回爬取结果字典
        result_dic['risk_level'] = TYPE_KPT_MAP[TYPE_KPT]  # 风险等级
        if pei_i is not None:  # 赔标 [存在 `赔`图标 即视为`赔标`]
            result_dic['pei'] = '赔标'
        else:
            result_dic['pei'] = '--'
        result_dic['progress'] = progress  # 完成度
        result_dic['user_name'] = user_name  # 用户名
        result_dic['credit_level'] = credit_level  # 信用等级
        result_dic['amount'] = new_lend_detail_money_list[0].get_text()  # 贷款金额
        result_dic['rate'] = new_lend_detail_money_list[1].get_text()  # 协议利率
        result_dic['term'] = new_lend_detail_money_list[2].get_text()  # 期限
        result_dic['sex'] = sex  # 性别
        result_dic['age'] = age  # 年龄
        result_dic['regist_date'] = regist_date  # 注册时间
        result_dic['edu_bg'] = edu_bg  # 文化程度
        result_dic['graduate'] = graduate  # 毕业院校
        result_dic['learn_way'] = learn_way  # 学习形式
        result_dic['lend_purpose'] = lend_purpose  # 借款用途
        result_dic['payment'] = payment  # 还款来源
        result_dic['work_info'] = work_info  # 工作信息
        result_dic['income_info'] = income_info  # 收入情况
        result_dic['verfied_info'] = verfied_info  # 认证信息
        # Fixed: 完成页面不包含统计信息
        temp = statistics_info_list[0]
        result_dic['sucuess_cnt'] = "-" if (not temp) else temp.get_text().strip()  # 成功借款次数
        temp = statistics_info_list[1]
        result_dic['first_sucuess_date'] = "-" if (not temp) else temp.get_text().strip()  # 第一次成功借款时间
        temp = statistics_info_list[2]
        result_dic['history_info'] = "-" if (not temp) else temp.get_text().strip()  # 历史记录
        temp = statistics_info_list[3]
        result_dic['sucuess_repayment_cnt'] = "-" if (not temp) else temp.get_text().strip()  # 成功还款次数
        temp = statistics_info_list[4]
        result_dic['normal_repayment_cnt'] = "-" if (not temp) else temp.get_text().strip()  # 正常还清次数
        temp = statistics_info_list[5]
        result_dic['delay_lt15_repayment_cnt'] = "-" if (not temp) else temp.get_text().strip()  # 逾期(0-15天)还清次数
        temp = statistics_info_list[6]
        result_dic['delay_gt15_repayment_cnt'] = "-" if (not temp) else temp.get_text().strip()  # 逾期(15天以上)还清次数
        temp = statistics_info_list[7]
        result_dic['amount_sum'] = "-" if (not temp) else temp.get_text().strip().replace('¥', '')  # 累计借款金额
        temp = statistics_info_list[8]
        result_dic['unreturned_amount'] = "-" if (not temp) else temp.get_text().strip().replace('¥', '')  # 待还金额
        temp = statistics_info_list[9]
        result_dic['unreceived_amount'] = "-" if (not temp) else temp.get_text().strip().replace('¥', '')  # 待收金额
        temp = statistics_info_list[10]
        result_dic['biggest_lend_amount'] = "-" if (not temp) else temp.get_text().strip().replace('¥', '')  # 单笔最高借款金额
        temp = statistics_info_list[11]
        result_dic['biggest_debt_amount'] = "-" if (not temp) else temp.get_text().strip().replace('¥', '')  # 历史最高负债
        result_dic['investor_list'] = investor_list  # 投资信息列表
        return result_dic
    except Exception as e:
        print("未知错误,链接地址：" + str(details_url) + "\n错误信息:")
        for info in e.args:  # 打印所有异常参数
            print(info)
        return {}  # 返回不包含任何信息的集合


# 获取总页数 [用于修正]
def total_page_getter(url):
    soup = html_to_soup(url)
    if '很抱歉，热门列表已经被抢空啦' in str(soup):  # 页面不包含列表信息
        total_page = -1
    else:
        page_info = soup.find('span', attrs={'class': 'pagerstatus'}).replace('共', '').replace('页', '').strip()
        if page_info is None:  # 可能出现不存在的情况
            total_page = -1
        else:
            total_page = int(page_info, base=10)
    return total_page


# 爬取逻辑
def data_spider(total_page=100):
    current_page = 1
    data_list = []
    while current_page <= total_page:
        url = url_constructor(current_page, TYPE_KPT)
        details_url_list = details_url_list_getter(url)
        if len(details_url_list) > 0:
            print('爬取第' + str(current_page) + '页数据。')
            for it in details_url_list:
                url = 'https:' + it
                print('爬取链接:' + url)
                data_list.append(details_info_getter(url))
                # 网站有一定的反爬虫机制，长期规律性连接可能导致Ip被锁或黑名单
                sleep_second = random.randint(0, 1)
                print('随机休眠' + str(sleep_second) + '秒...')
                time.sleep(sleep_second)
            data_output_xls(data_list)  # 输出数据
        else:
            print('Error:未找到数据!\n尝试重新获取...')
            current_page = 0
            total_page = total_page_getter(url)
            if total_page == -1:  # 如果返回-1 则表示当前没有数据,休眠后继续尝试
                time.sleep(10)
                total_page = 100
        current_page += 1


# 输出数据到excel
def data_output_xls(data_list):
    print('数据写入文件开始....')
    xls_data_column_length = len(data_xls_poz_map)
    # 文件不存在则创建文件
    if not os.path.exists(file_name):
        wb = Workbook()
    else:
        wb = load_workbook(file_name)
    title = "拍拍贷数据" + str(today) + '_' + TYPE_KPT_MAP[TYPE_KPT]
    if title not in wb.sheetnames:
        work_sheet = wb.create_sheet(title=title)
    else:
        work_sheet = wb[title]
    # 标题行
    for i in range(1, xls_data_column_length):
        _ = work_sheet.cell(column=i, row=1, value="%s" % data_xls_poz_map[i]['name'])
    # 数据行
    global row
    for it in data_list:
        investor_list_size = len(it['investor_list'])  # 投资人信息数量
        for i in range(1, xls_data_column_length):  # 最后两列不输出数据
            _ = work_sheet.cell(column=i, row=row, value="%s" % it[data_xls_poz_map[i]['code']])
        total = 0.00
        for i in range(investor_list_size):
            if '自动投标' == it['investor_list'][i]['investment_way']:  # 只输出自动投标
                total += float(it['investor_list'][i]['valid_amount'])  # 计算自动投标总和
        if investor_list_size > 0:
            _ = work_sheet.cell(column=29, row=row, value="%s" % '自动投标')  # 投资方式
            _ = work_sheet.cell(column=30, row=row, value="%s" % total)  # 自动投资总和
        row += 1
    try:
        wb.save(filename=file_name)
    except IOError as iox:
        print('文件读写异常')
        print(data_list)  # 将数据输出,避免数据因异常丢失
        print('错误信息:')
        for e in iox.args:
            print(e)
        print('数据写入文件失败....')
    except Exception as unknown:
        print("未知异常导致文件输出失败!错误信息:")
        for e in unknown.args:
            print(e)
    else:
        print('数据写入文件完成....')


# 人工登陆
def login():
    browser.get(login_url)
    if input('完成登陆后请输入任意字符'):
        return False
    else:
        return True


# Main method
if __name__ == '__main__':
    try:
        while login():
            print('等待登陆')
        '''
        循环爬取前两页数据
        如不需要循环爬取前两页数据只需要,删掉循环 
        使用 data_spider() 即可
        '''
        while True:
            data_spider(2)
    finally:
        browser.close()
